"""
Excel 结构查看器
用于快速查看 Excel 文件的 Sheet 列表和各 Sheet 的列名及少量示例数据，
帮助对齐检查脚本中的字段引用。

用法:
    python scripts/explorer.py <excel_path>
"""
import sys
from pathlib import Path

import openpyxl

# 每个 Sheet 最多展示的前 N 行示例数据
MAX_SAMPLE_ROWS = 10


def explore_excel(excel_path: str | Path) -> None:
    path = Path(excel_path)
    if not path.exists():
        print(f"[错误] 文件不存在: {path}")
        return

    print(f"=" * 60)
    print(f"Excel 结构: {path}")
    print(f"=" * 60)

    # 使用 openpyxl 获取所有 Sheet
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"\nSheet 列表 ({len(sheet_names)} 个):")
    for i, name in enumerate(sheet_names, 1):
        print(f"  {i}. {name}")

    # 读取每个 Sheet 的列名和少量示例行
    print()
    for sheet_name in sheet_names:
        try:
            ws = wb[sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"--- Sheet: {sheet_name} ---")
            print(f"  行数: {max_row}, 列数: {max_col}")

            # 读取前几行作为示例
            sample_end = min(MAX_SAMPLE_ROWS, max_row)
            sample_rows = list(ws.iter_rows(min_row=1, max_row=sample_end,
                                            min_col=1, max_col=max_col,
                                            values_only=True))
            for row_idx, row in enumerate(sample_rows, 1):
                cells = [str(cell) if cell is not None else '' for cell in row]
                print(f"  Row {row_idx}: {' | '.join(cells)}")
            if max_row > MAX_SAMPLE_ROWS:
                print(f"  ... 共 {max_row} 行，仅展示前 {MAX_SAMPLE_ROWS} 行示例")
            print()
        except Exception as e:
            print(f"--- Sheet: {sheet_name} ---")
            print(f"  [读取失败] {e}")
            print()

    wb.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python scripts/explorer.py <excel_path>")
        sys.exit(1)
    explore_excel(sys.argv[1])
