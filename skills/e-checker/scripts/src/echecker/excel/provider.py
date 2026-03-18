"""Excel数据提供器"""

from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import openpyxl
from openpyxl.cell.cell import Cell

from echecker.excel.cell_ref import CellRef, CellRange, parse_cell_ref


class ExcelProvider:
    """Excel数据提供器 - 负责读取Excel文件内容"""

    def __init__(self, path: Union[str, Path]):
        self.path = Path(path) if isinstance(path, str) else path
        self._workbook: Optional[openpyxl.Workbook] = None
        self._cache: Dict[str, Any] = {}

    def open(self) -> "ExcelProvider":
        """打开Excel文件"""
        if not self.path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {self.path}")

        self._workbook = openpyxl.load_workbook(self.path, data_only=True)
        return self

    def close(self) -> None:
        """关闭Excel文件"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None

    def __enter__(self) -> "ExcelProvider":
        return self.open()

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

    def get_sheet_names(self) -> List[str]:
        """获取所有Sheet名称"""
        if not self._workbook:
            raise RuntimeError("Excel文件未打开")
        return self._workbook.sheetnames

    def get_cell_value(self, ref: Union[str, CellRef]) -> Any:
        """获取单元格值"""
        if isinstance(ref, str):
            ref = CellRef.from_string(ref)

        cache_key = f"{ref.sheet}.{ref.row}.{ref.col}"
        if cache_key in self._cache:
            return self._cache[cache_key]

        if not self._workbook:
            raise RuntimeError("Excel文件未打开")

        if ref.sheet not in self._workbook.sheetnames:
            raise ValueError(f"Sheet不存在: {ref.sheet}")

        sheet = self._workbook[ref.sheet]
        cell = sheet.cell(row=ref.row, column=ref.col)
        value = cell.value

        self._cache[cache_key] = value
        return value

    def get_range_values(self, ref: Union[str, CellRange]) -> Dict[str, Any]:
        """获取范围值"""
        if isinstance(ref, str):
            ref = CellRange.from_string(ref)

        if not self._workbook:
            raise RuntimeError("Excel文件未打开")

        if ref.sheet not in self._workbook.sheetnames:
            raise ValueError(f"Sheet不存在: {ref.sheet}")

        sheet = self._workbook[ref.sheet]
        values = {}

        for row in range(ref.start_row, ref.end_row + 1):
            for col in range(ref.start_col, ref.end_col + 1):
                cell_ref = f"{ref.sheet}.{CellRef._col_to_letter(col)}{row}"
                cell = sheet.cell(row=row, column=col)
                values[cell_ref] = cell.value
                self._cache[f"{ref.sheet}.{row}.{col}"] = cell.value

        return values

    def get_column_values(self, sheet: str, col: Union[int, str]) -> List[Any]:
        """获取整列值"""
        if not self._workbook:
            raise RuntimeError("Excel文件未打开")

        if sheet not in self._workbook.sheetnames:
            raise ValueError(f"Sheet不存在: {sheet}")

        ws = self._workbook[sheet]

        if isinstance(col, str):
            col = CellRef._letter_to_col(col)

        values = []
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            values.append(cell.value)

        return values

    def get_row_values(self, sheet: str, row: int) -> List[Any]:
        """获取整行值"""
        if not self._workbook:
            raise RuntimeError("Excel文件未打开")

        if sheet not in self._workbook.sheetnames:
            raise ValueError(f"Sheet不存在: {sheet}")

        ws = self._workbook[sheet]
        values = []

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            values.append(cell.value)

        return values

    def find_cells(self, sheet: str, predicate) -> List[CellRef]:
        """根据条件查找单元格"""
        if not self._workbook:
            raise RuntimeError("Excel文件未打开")

        if sheet not in self._workbook.sheetnames:
            raise ValueError(f"Sheet不存在: {sheet}")

        ws = self._workbook[sheet]
        results = []

        for row in ws.iter_rows():
            for cell in row:
                if predicate(cell.value):
                    results.append(CellRef(
                        sheet=sheet,
                        row=cell.row,
                        col=cell.column
                    ))

        return results

    def get_sheet_dimensions(self, sheet: str) -> tuple:
        """获取Sheet维度 (max_row, max_col)"""
        if not self._workbook:
            raise RuntimeError("Excel文件未打开")

        if sheet not in self._workbook.sheetnames:
            raise ValueError(f"Sheet不存在: {sheet}")

        ws = self._workbook[sheet]
        return (ws.max_row, ws.max_column)
