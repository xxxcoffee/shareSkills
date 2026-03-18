"""Excel标注器 - 在Excel中标注错误单元格"""

from pathlib import Path
from typing import Union, List, Dict

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment

from echecker.types import ValidationError, Severity


class ExcelAnnotator:
    """Excel标注器 - 在原始Excel文件中标注错误"""

    ERROR_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    ERROR_FONT = Font(color="FFFFFF", bold=True)

    def __init__(self, source_path: Union[str, Path]):
        self.source_path = Path(source_path)
        self._workbook: openpyxl.Workbook = None

    def open(self) -> "ExcelAnnotator":
        """打开Excel文件"""
        self._workbook = openpyxl.load_workbook(self.source_path)
        return self

    def close(self) -> None:
        """关闭Excel文件"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None

    def __enter__(self) -> "ExcelAnnotator":
        return self.open()

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

    def annotate_errors(self, errors: List[ValidationError]) -> None:
        """标注错误到Excel文件"""
        # 按Sheet和单元格分组
        errors_by_cell: Dict[tuple, List[ValidationError]] = {}
        for error in errors:
            # 解析单元格地址 (Sheet1.A1 -> Sheet1, A1)
            parts = error.cell_ref.split('.')
            if len(parts) == 2:
                sheet_name, cell_addr = parts
            else:
                sheet_name = error.sheet_name
                cell_addr = error.cell_ref
            key = (sheet_name, cell_addr)
            if key not in errors_by_cell:
                errors_by_cell[key] = []
            errors_by_cell[key].append(error)

        # 标注每个单元格
        for (sheet_name, cell_addr), cell_errors in errors_by_cell.items():
            if sheet_name not in self._workbook.sheetnames:
                continue

            ws = self._workbook[sheet_name]

            # 确定标注样式
            has_error = any(e.severity == Severity.ERROR for e in cell_errors)
            fill = self.ERROR_FILL if has_error else self.WARNING_FILL

            try:
                cell = ws[cell_addr]
                cell.fill = fill

                # 添加注释
                comments = [f"[{e.severity.value.upper()}] {e.message}" for e in cell_errors]
                comment_text = "\n".join(comments)
                cell.comment = Comment(comment_text, "eChecker")
            except Exception:
                pass  # 忽略无法标注的单元格

    def save(self, output_path: Union[str, Path] = None) -> Path:
        """保存标注后的文件"""
        if output_path is None:
            # 在原文件名后添加 _annotated
            stem = self.source_path.stem
            suffix = self.source_path.suffix
            output_path = self.source_path.parent / f"{stem}_annotated{suffix}"

        self._workbook.save(output_path)
        return Path(output_path)
