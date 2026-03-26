"""文件夹通配符展开器

支持通配符语法批量展开规则：
- 通配符语法：folder/*.xlsx:*.range（匹配文件夹下所有xlsx文件的所有Sheet）
- 递归语法：folder/**/*.xlsx:*.range（递归子文件夹）
- Sheet通配符：folder/*.xlsx:*.range（*匹配所有Sheet）
"""

import fnmatch
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, List, Optional

if TYPE_CHECKING:
    from .v3_parser import V3Rule, V3ValidationConfig


@dataclass
class ExpandedTarget:
    """展开后的目标"""
    file_path: str
    sheet_name: str
    range_str: str
    original_target: str


@dataclass
class ParsedTarget:
    """解析后的目标组件"""
    file_pattern: str
    sheet_pattern: str
    range_str: str
    is_recursive: bool = False


class FolderExpander:
    """文件夹通配符展开器

    将包含通配符的 target 展开为多个具体的 target，
    支持文件路径通配符和 Sheet 名称通配符。

    示例：
        folder/*.xlsx:*.A1:*  -> 展开为 folder/file1.xlsx:Sheet1.A1:*, folder/file1.xlsx:Sheet2.A1:*, ...
        folder/**/*.xlsx:Data.A1:B10  -> 递归匹配所有子文件夹中的 xlsx 文件
    """

    def __init__(self, base_path: Optional[str] = None):
        """初始化展开器

        Args:
            base_path: 基础路径，默认为当前目录
        """
        self.base_path = Path(base_path) if base_path else Path(".")

    def is_wildcard_target(self, target: str) -> bool:
        """判断 target 是否包含通配符

        Args:
            target: 目标字符串，格式为 "file_pattern:Sheet.range"

        Returns:
            如果包含通配符返回 True，否则返回 False
        """
        try:
            parsed = self._parse_target(target)
            return self._has_wildcard(parsed)
        except ValueError:
            return False

    def _has_wildcard(self, parsed: ParsedTarget) -> bool:
        """检查解析后的目标是否包含通配符"""
        return (
            "*" in parsed.file_pattern
            or "?" in parsed.file_pattern
            or "*" in parsed.sheet_pattern
            or "?" in parsed.sheet_pattern
        )

    def _parse_target(self, target: str) -> ParsedTarget:
        """解析 target 字符串

        支持格式：
        - file_pattern:Sheet.range
        - file_pattern:Sheet:range（兼容格式）

        Args:
            target: 目标字符串

        Returns:
            ParsedTarget 对象

        Raises:
            ValueError: 如果格式无效
        """
        # 分离文件部分和 Sheet/Range 部分
        if ":" not in target:
            raise ValueError(f"Invalid target format: {target}. Expected 'file:Sheet.range' or 'file:Sheet:range'")

        # 找到第一个冒号，分离文件路径和后面的部分
        colon_idx = target.find(":")
        file_pattern = target[:colon_idx]
        rest = target[colon_idx + 1:]

        # 解析 Sheet 和 Range 部分
        # 支持两种格式：Sheet.range 或 Sheet:range
        sheet_pattern = ""
        range_str = ""

        if "." in rest:
            # 格式：Sheet.range
            # 需要找到最后一个点，因为 Sheet 名可能包含括号
            # 例如：element(PassNew).A1:*
            # 从右向左找第一个点
            dot_idx = rest.rfind(".")
            sheet_pattern = rest[:dot_idx]
            range_str = rest[dot_idx + 1:]
        elif ":" in rest:
            # 格式：Sheet:range
            colon_idx2 = rest.find(":")
            sheet_pattern = rest[:colon_idx2]
            range_str = rest[colon_idx2 + 1:]
        else:
            raise ValueError(f"Invalid target format: {target}. Cannot parse Sheet and range.")

        # 检查是否递归模式
        is_recursive = "**" in file_pattern

        return ParsedTarget(
            file_pattern=file_pattern,
            sheet_pattern=sheet_pattern,
            range_str=range_str,
            is_recursive=is_recursive
        )

    def expand(self, rules: List["V3Rule"]) -> List["V3Rule"]:
        """展开所有包含通配符的规则

        Args:
            rules: 原始规则列表

        Returns:
            展开后的规则列表（不包含通配符的具体规则）
        """
        expanded_rules = []

        for rule in rules:
            if not self.is_wildcard_target(rule.target):
                # 不包含通配符，直接保留
                expanded_rules.append(rule)
                continue

            # 展开通配符规则
            expanded_targets = self._expand_target(rule.target)

            for idx, expanded_target in enumerate(expanded_targets):
                # 创建新的规则，保持原有规则的所有字段
                new_rule = self._create_expanded_rule(rule, expanded_target, idx)
                expanded_rules.append(new_rule)

        return expanded_rules

    def _expand_target(self, target: str) -> List[ExpandedTarget]:
        """展开单个通配符 target

        Args:
            target: 包含通配符的 target 字符串

        Returns:
            展开后的 ExpandedTarget 列表
        """
        parsed = self._parse_target(target)

        # 匹配文件
        matched_files = self._match_files(parsed.file_pattern)

        expanded_targets = []
        for file_path in matched_files:
            # 匹配 Sheet
            matched_sheets = self._match_sheets(file_path, parsed.sheet_pattern)

            for sheet_name in matched_sheets:
                expanded_targets.append(ExpandedTarget(
                    file_path=str(file_path),
                    sheet_name=sheet_name,
                    range_str=parsed.range_str,
                    original_target=target
                ))

        return expanded_targets

    def _match_files(self, file_pattern: str) -> List[Path]:
        """根据文件模式匹配文件

        Args:
            file_pattern: 文件模式，如 "folder/*.xlsx" 或 "folder/**/*.xlsx"

        Returns:
            匹配的文件路径列表
        """
        # 解析路径
        if "**" in file_pattern:
            # 递归模式
            parts = file_pattern.split("**")
            base_dir = self.base_path / parts[0].rstrip("/")
            rest_pattern = parts[1].lstrip("/") if len(parts) > 1 else ""

            if not base_dir.exists():
                return []

            matched_files = []
            for file_path in base_dir.rglob(rest_pattern or "*"):
                if file_path.is_file() and self._match_pattern(str(file_path.relative_to(self.base_path)), file_pattern):
                    matched_files.append(file_path)
            return matched_files
        else:
            # 非递归模式
            # 找到基础目录（通配符之前的部分）
            pattern_parts = file_pattern.split("/")
            base_parts = []
            pattern_start_idx = 0

            for i, part in enumerate(pattern_parts):
                if "*" in part or "?" in part:
                    pattern_start_idx = i
                    break
                base_parts.append(part)

            if base_parts:
                base_dir = self.base_path / "/".join(base_parts)
            else:
                base_dir = self.base_path

            if not base_dir.exists():
                return []

            # 构建相对模式
            rel_pattern = "/".join(pattern_parts[pattern_start_idx:])

            matched_files = []
            search_dir = base_dir if base_parts else self.base_path

            for file_path in search_dir.glob(rel_pattern):
                if file_path.is_file():
                    matched_files.append(file_path)

            return matched_files

    def _match_pattern(self, path: str, pattern: str) -> bool:
        """使用 fnmatch 匹配路径，支持 ** 递归匹配

        ** 匹配规则:
        - **/ 匹配零层或多层目录前缀
        - /**/ 匹配中间任意层目录
        - /** 匹配任意后缀
        """
        # 处理 ** 模式：将 ** 转换为匹配任意路径的等价形式
        if "**" in pattern:
            import re

            # 使用不会冲突的唯一占位符
            PREFIX_PLACEHOLDER = "__GLOBSTAR_PREFIX__"
            SUFFIX_PLACEHOLDER = "__GLOBSTAR_SUFFIX__"
            MIDDLE_PLACEHOLDER = "__GLOBSTAR_MIDDLE__"
            GLOBSTAR_PLACEHOLDER = "__GLOBSTAR__"

            regex_pattern = pattern

            # 1. 处理 /**/ 中间模式：匹配 / 加上任意中间路径
            regex_pattern = regex_pattern.replace("/**/", MIDDLE_PLACEHOLDER)

            # 2. 处理开头的 **/
            if regex_pattern.startswith("**/"):
                regex_pattern = PREFIX_PLACEHOLDER + regex_pattern[3:]

            # 3. 处理结尾的 /**
            if regex_pattern.endswith("/**"):
                regex_pattern = regex_pattern[:-3] + SUFFIX_PLACEHOLDER

            # 4. 处理剩余的 **（前后都不是 /）
            regex_pattern = regex_pattern.replace("**", GLOBSTAR_PLACEHOLDER)

            # 5. 转义正则特殊字符（注意：占位符使用下划线，不会被转义）
            special_chars = r".^$+{}[]|()"
            for char in special_chars:
                regex_pattern = regex_pattern.replace(char, f"\\{char}")

            # 6. 处理单 *：匹配任意字符（非路径分隔符）
            regex_pattern = regex_pattern.replace("*", "[^/]*")

            # 7. 处理 ?：匹配单个字符
            regex_pattern = regex_pattern.replace("?", "[^/]")

            # 8. 恢复 ** 的各种形式
            regex_pattern = regex_pattern.replace(PREFIX_PLACEHOLDER, "(?:.*/)?")
            regex_pattern = regex_pattern.replace(SUFFIX_PLACEHOLDER, "(?:/.*)?")
            regex_pattern = regex_pattern.replace(MIDDLE_PLACEHOLDER, "(?:/.*)?/")
            regex_pattern = regex_pattern.replace(GLOBSTAR_PLACEHOLDER, ".*")

            # 添加锚点
            regex_pattern = f"^{regex_pattern}$"

            return bool(re.match(regex_pattern, path))
        else:
            return fnmatch.fnmatch(path, pattern)

    def _match_sheets(self, file_path: Path, sheet_pattern: str) -> List[str]:
        """根据 Sheet 模式匹配 Sheet 名称

        Args:
            file_path: Excel 文件路径
            sheet_pattern: Sheet 名称模式，如 "*" 或 "Data*"

        Returns:
            匹配的 Sheet 名称列表
        """
        # 如果没有通配符，直接返回原名称
        if "*" not in sheet_pattern and "?" not in sheet_pattern:
            return [sheet_pattern]

        # 读取 Excel 文件获取所有 Sheet 名称
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception:
            # 如果无法读取文件，返回空列表
            return []

        # 使用 fnmatch 匹配 Sheet 名称
        matched_sheets = [
            name for name in sheet_names
            if fnmatch.fnmatch(name, sheet_pattern)
        ]

        return matched_sheets

    def _create_expanded_rule(
        self,
        original_rule: "V3Rule",
        expanded_target: ExpandedTarget,
        index: int
    ) -> "V3Rule":
        """创建展开后的规则

        Args:
            original_rule: 原始规则
            expanded_target: 展开后的目标
            index: 展开索引，用于生成唯一 ID

        Returns:
            新的 V3Rule 对象
        """
        # 构建新的 target 字符串
        # 格式：file_path:Sheet.range
        new_target = f"{expanded_target.file_path}:{expanded_target.sheet_name}.{expanded_target.range_str}"

        # 生成新的 ID
        new_id = original_rule.id
        if new_id:
            new_id = f"{new_id}_{index}"
        else:
            # 如果没有原 ID，生成基于 target 的 ID
            import hashlib
            new_id = hashlib.md5(new_target.encode()).hexdigest()[:8]

        # 复制验证配置
        new_validations = []
        for val in original_rule.validations:
            # 创建验证配置的深拷贝
            # 延迟导入以避免循环导入
            from .v3_parser import V3ValidationConfig
            new_val = V3ValidationConfig(
                validation_type=val.validation_type,
                config=dict(val.config) if val.config else {},
                message=val.message,
                pipeline=val.pipeline,  # Pipeline 对象通常是不可变的，可以直接引用
                raw_config=dict(val.raw_config) if val.raw_config else {},
                has_templates=val.has_templates
            )
            new_validations.append(new_val)

        # 创建新规则
        # 延迟导入 V3Rule 以避免循环导入
        from .v3_parser import V3Rule
        return V3Rule(
            target=new_target,
            validations=new_validations,
            id=new_id,
            description=original_rule.description,
            enabled=original_rule.enabled
        )

    def expand_target_string(self, target: str) -> List[str]:
        """展开 target 字符串为具体的目标字符串列表

        这是一个便捷方法，用于直接展开 target 而不需要完整的规则对象。

        Args:
            target: 包含通配符的 target 字符串

        Returns:
            展开后的 target 字符串列表
        """
        expanded = self._expand_target(target)
        return [
            f"{et.file_path}:{et.sheet_name}.{et.range_str}"
            for et in expanded
        ]
