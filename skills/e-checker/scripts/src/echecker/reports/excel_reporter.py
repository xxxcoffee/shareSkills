"""Excel报告生成器"""

from pathlib import Path
from typing import Union

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from echecker.types import ValidationReport, Severity
from echecker.reports.base import BaseReporter


class ExcelReporter(BaseReporter):
    """Excel标注报告生成器"""

    ERROR_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")

    def __init__(self, source_excel: Union[str, Path] = None):
        self.source_excel = Path(source_excel) if source_excel else None

    def generate(self, report: ValidationReport, output_path: Union[str, Path] = None) -> str:
        """生成Excel标注报告"""
        output_path = Path(output_path) if output_path else Path("validation_report.xlsx")

        if self.source_excel and self.source_excel.exists():
            wb = openpyxl.load_workbook(self.source_excel)
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # 创建错误摘要Sheet
        self._create_summary_sheet(wb, report)

        # 标注错误单元格
        self._annotate_errors(wb, report)

        wb.save(output_path)
        return str(output_path)

    def _create_summary_sheet(self, wb: openpyxl.Workbook, report: ValidationReport) -> None:
        """创建错误摘要Sheet"""
        if "ValidationReport" in wb.sheetnames:
            wb.remove(wb["ValidationReport"])

        ws = wb.create_sheet("ValidationReport", 0)

        # 标题
        ws.append(["Excel配置检查报告"])
        ws.append([])

        # 摘要
        summary = report.summary
        ws.append(["摘要统计"])
        ws.append(["总规则数", summary.total_rules])
        ws.append(["校验单元格", summary.total_cells_checked])
        ws.append(["通过数", summary.passed_count])
        ws.append(["错误数", summary.error_count])
        ws.append(["警告数", summary.warning_count])
        ws.append(["耗时(秒)", round(summary.duration_seconds, 2)])
        ws.append([])

        # 错误详情
        if report.errors:
            ws.append(["错误详情"])
            ws.append(["规则ID", "Sheet", "单元格", "错误类型", "严重程度", "消息", "期望值", "实际值"])

            for error in report.errors:
                ws.append([
                    error.rule_id,
                    error.sheet_name,
                    error.cell_ref,
                    error.error_type.name if hasattr(error.error_type, 'name') else str(error.error_type),
                    error.severity.value,
                    error.message,
                    str(error.expected) if error.expected is not None else "",
                    str(error.actual) if error.actual is not None else ""
                ])

        # 设置列宽
        for col_idx in range(1, 9):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    def _annotate_errors(self, wb: openpyxl.Workbook, report: ValidationReport) -> None:
        """在原始Sheet中标注错误单元格"""
        errors_by_cell = {}
        for error in report.errors:
            key = (error.sheet_name, error.cell_ref.split(".")[-1])  # (Sheet名, 单元格地址)
            if key not in errors_by_cell:
                errors_by_cell[key] = []
            errors_by_cell[key].append(error)

        for (sheet_name, cell_addr), errors in errors_by_cell.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # 获取最高严重级别
            has_error = any(e.severity == Severity.ERROR for e in errors)
            fill = self.ERROR_FILL if has_error else self.WARNING_FILL

            try:
                cell = ws[cell_addr]
                cell.fill = fill

                # 添加注释
                comment_text = "; ".join(e.message for e in errors)
                from openpyxl.comments import Comment
                cell.comment = Comment(comment_text, "eChecker")
            except Exception:
                pass  # 忽略无法标注的单元格
